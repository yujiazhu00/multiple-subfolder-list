import os
from pathlib import Path
import openpyxl

def create_master_list(my_dir):
    p = Path(my_dir)
    return os.listdir(p)

def extract_date(list):
    date_list = []
    len_list = len(list)
    for i in range(0, len_list):
        date_list = date_list + [list[i][0:10]]
    return date_list

def extract_name(list):
    name_list = []
    len_list = len(list)
    for i in range(0, len_list):
        max_word = len(list[i]) - 4
        if max_word < 11:
            name_list = name_list + [list[i][10:max_word]] #only to be used if there are short file names - and if so, change 'if' on the next line to 'elif'
        elif list[i][10] == " ":
            name_list = name_list + [list[i][11:max_word]]
        else:
            name_list = name_list + [list[i][10:max_word]]
    return name_list

def save_excel(date_list, name_list,name_output_file):
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    length_list = len(date_list)
    sheet.cell(row=1, column=1).value = 'S/No.'
    sheet.cell(row=1, column=2).value = 'Date of Document (YYYY.MM.DD)'
    sheet.cell(row=1, column=3).value = 'File Name'
    for i in range(1, length_list + 1):
        sheet.cell(row=i+1, column=1).value = str(i) + '.'
        sheet.cell(row=i+1, column=2).value = date_list[i - 1]
        sheet.cell(row=i+1, column=3).value = name_list[i - 1]
    wb.save(name_output_file)

def list_of_document(input_dir, output_name):
    master_dir = create_master_list(input_dir)
    my_date_list = extract_date(master_dir)
    my_name_list = extract_name(master_dir)
    save_excel(my_date_list, my_name_list, output_name)

def multiple_folders_LOD(location):
    p = Path(location)
    list_folders = os.listdir(p)
    len_list = len(list_folders)
    for i in range(0,len_list):
        path_interest = p/list_folders[i]
        file_name = list_folders[i] + '.xlsx'
        list_of_document(path_interest, file_name)
